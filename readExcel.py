from openpyxl import load_workbook


#------------Create the Dictionary for Business Need to Deliverables and BusNeed to Solutions-----------------
wb = load_workbook(filename='businessNeed.xlsx', read_only=True)
ws = wb['owssvr']

rows = []
for row in ws.rows:
    r = []
    for cell in row:
        r.append(cell.value)
    rows.append(r)


rows = rows[2:]

businessNeeds = [cell[0] for cell in rows]
delivrablesColumn = [cell[2] for cell in rows]

deliverable= []

for d in delivrablesColumn:
    output = []
    elementsInCol = d.split(";#")
    elementsInCol.remove("_Starting Over")
    for item in elementsInCol:
        if item.isdigit() == False:
            output.append(item)
    deliverable.append(output)


solutionsColumn = [cell[1] for cell in rows]

solutions= []

for d in solutionsColumn:
    output = []
    elementsInCol = d.split(";#")
    elementsInCol.remove("_Starting Over")
    for item in elementsInCol:
        if item.isdigit() == False:
            output.append(item)
    solutions.append(output)

busNeedToDeliverable = {}

for i in range(len(businessNeeds)):
    value = []
    if len(deliverable[i])==0: value = ["NotDefined"]
    else: value = deliverable[i]
    busNeedToDeliverable[businessNeeds[i]]=value

busNeedToSolution = {}
for i in range(len(businessNeeds)):
    value = []
    if len(solutions[i])==0: value = ["NotDefined"]
    else: value = solutions[i]
    busNeedToSolution[businessNeeds[i]]=value





#------------Create the Dictionary for Deliverables to Solution, and Deliverables to BusNeed-----------------
wb = load_workbook(filename='deliverables.xlsx', read_only=True)
ws = wb['owssvr']

rows = []
for row in ws.rows:
    r = []
    for cell in row:
        r.append(cell.value)
    rows.append(r)


rows = rows[2:]

deliverable = [cell[0] for cell in rows]
relatedSoftwareCol = [cell[2] for cell in rows]

relatedSoftware= []

for d in relatedSoftwareCol:
    output = []
    elementsInCol = d.split(";#")
    elementsInCol.remove("_Starting Over")
    for item in elementsInCol:
        if item.isdigit() == False:
            output.append(item)
    relatedSoftware.append(output)


relatedBusNeedCol = [cell[1] for cell in rows]

relatedBusNeed= []

for d in relatedBusNeedCol:
    output = []
    elementsInCol = d.split(";#")
    elementsInCol.remove("_Starting Over")
    for item in elementsInCol:
        if item.isdigit() == False:
            output.append(item)
    relatedBusNeed.append(output)


deliverablesToSolution = {}

for i in range(len(deliverable)):
    value = []
    if len(relatedSoftware[i])==0: value = ["NotDefined"]
    else: value = relatedSoftware[i]
    deliverablesToSolution[deliverable[i]]=value

deliverablesToBusNeed = {}

for i in range(len(deliverable)):
    value = []
    if len(relatedBusNeed[i])==0: value = ["NotDefined"]
    else: value = relatedBusNeed[i]
    deliverablesToBusNeed[deliverable[i]]=value



#------------Create the Dictionary for Solution to Deliverable, and solution to BusNeed-----------------
wb = load_workbook(filename='solutions.xlsx', read_only=True)
ws = wb['owssvr']

rows = []
for row in ws.rows:
    r = []
    for cell in row:
        r.append(cell.value)
    rows.append(r)


rows = rows[2:]

solutions = [cell[0] for cell in rows]
relatedDeliverablesCol = [cell[2] for cell in rows]

relatedDel= []

for d in relatedDeliverablesCol:
    output = []
    elementsInCol = d.split(";#")
    elementsInCol.remove("_Starting Over")
    for item in elementsInCol:
        if item.isdigit() == False:
            output.append(item)
    relatedDel.append(output)


relatedBusNeedCol = [cell[1] for cell in rows]

relatedBusNeed= []

for d in relatedBusNeedCol:
    output = []
    elementsInCol = d.split(";#")
    elementsInCol.remove("_Starting Over")
    for item in elementsInCol:
        if item.isdigit() == False:
            output.append(item)
    relatedBusNeed.append(output)


solutionToDel = {}

for i in range(len(solutions)):
    value = []
    if len(relatedDel[i])==0: value = ["NotDefined"]
    else: value = relatedDel[i]
    solutionToDel[solutions[i]]=value

solutionToNeed = {}

for i in range(len(solutions)):
    value = []
    if len(relatedBusNeed[i])==0: value = ["NotDefined"]
    else: value = relatedBusNeed[i]
    solutionToNeed[solutions[i]]=value
